import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import os
from openpyxl import Workbook

def create_default_excel(filename):
    if not os.path.exists(filename):
        print(f"Creating new file: {filename}")
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        
        # Define the headers
        headers = ["Roll No", "Student Email", "CI", "Python", "Data Mining"]
        sheet.append(headers)
        
        # Adding some sample data so the code doesn't crash on empty rows
        # Format: [Roll, Email, CI_Absences, Python_Absences, DM_Absences]
        sample_students = [
            [101, "student1@gmail.com", 0, 0, 0],
            [102, "student2@gmail.com", 0, 0, 0],
        ]
        for student in sample_students:
            sheet.append(student)
            
        wb.save(filename)
        print("Excel file created successfully!")

# Use a relative path so it stays inside your project folder
FILE_NAME = "attendance.xlsx"
create_default_excel(FILE_NAME)


# Constants
FILE_PATH = 'attendance.xlsx'
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "your_email@gmail.com"
SENDER_PWD = "your_app_password" # Use App Password!

def send_emails(recipients, subject, body):
    """Refactored email function to handle multiple recipients efficiently"""
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PWD)
        
        for to_id in recipients:
            message = MIMEMultipart()
            message['Subject'] = subject
            message.attach(MIMEText(body, 'plain'))
            server.sendmail(SENDER_EMAIL, to_id, message.as_string())
            
        server.quit()
        print(f"Emails sent successfully to {len(recipients)} people.")
    except Exception as e:
        print(f"Error sending email: {e}")

def main():
    try:
        book = openpyxl.load_workbook(FILE_PATH)
        sheet = book['Sheet1']
    except FileNotFoundError:
        print("Excel file not found!")
        return

    subjects = {1: "CI", 2: "Python", 3: "Data Mining"}
    staff_mails = ['staff1@gmail.com', 'staff2@gmail.com', 'staff3@gmail.com']

    while True:
        print("\n1--->CI | 2--->Python | 3--->DM")
        sub_choice = int(input("Enter subject number: "))
        if sub_choice not in subjects: break

        absent_rolls = input("Enter roll numbers separated by space: ").split()
        absent_rolls = [int(r) for r in absent_rolls]

        subject_name = subjects[sub_choice]
        col_idx = sub_choice + 2 # Assuming Roll=1, Email=2, CI=3, Py=4, DM=5
        
        lack_list_rolls = []
        lack_list_emails = []

        for row in range(2, sheet.max_row + 1):
            roll_in_sheet = sheet.cell(row=row, column=1).value
            
            if roll_in_sheet in absent_rolls:
                # Update attendance
                current_absences = sheet.cell(row=row, column=col_idx).value or 0
                new_absences = current_absences + 1
                sheet.cell(row=row, column=col_idx).value = new_absences
                
                student_email = sheet.cell(row=row, column=2).value

                # Threshold Logic
                if new_absences == 2:
                    msg = f"Warning! You can take only one more leave for {subject_name}."
                    send_emails([student_email], "Attendance Warning", msg)
                
                elif new_absences > 2:
                    lack_list_rolls.append(str(roll_in_sheet))
                    lack_list_emails.append(student_email)

        # Notify staff and students about critical lack of attendance
        if lack_list_rolls:
            staff_msg = f"Students with critical lack in {subject_name}: {', '.join(lack_list_rolls)}"
            stu_msg = f"You have a critical lack of attendance in {subject_name}!"
            
            send_emails([staff_mails[sub_choice-1]], "Lack of Attendance Report", staff_msg)
            send_emails(lack_list_emails, "Critical Attendance Alert", stu_msg)

        book.save(FILE_PATH)
        if input("Another subject? (y/n): ").lower() != 'y': break

if __name__ == "__main__":
    main()